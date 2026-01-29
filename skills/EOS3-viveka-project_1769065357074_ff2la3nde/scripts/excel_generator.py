# Excel台账生成模块
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from typing import List, Dict

class ExcelGenerator:
    def __init__(self, output_path: str = 'expense_report.xlsx'):
        self.output_path = output_path
        self.workbook = Workbook()
    
    def generate(self, invoice_data: List[Dict]):
        '''生成包含明细和汇总的Excel台账'''
        # 创建明细表
        self._create_detail_sheet(invoice_data)
        # 创建汇总表
        self._create_summary_sheet(invoice_data)
        # 保存文件
        self.workbook.save(self.output_path)
    
    def _create_detail_sheet(self, data: List[Dict]):
        '''创建明细工作表'''
        ws = self.workbook.active
        ws.title = '发票明细'
        
        # 设置表头
        headers = ['序号', '购买方名称', '开票日期', '发票类型', '发票号码', '金额', '税额', '价税合计']
        ws.append(headers)
        
        # 设置表头样式
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
        
        # 按日期排序
        sorted_data = sorted(data, key=lambda x: x.get('invoice_date', ''))
        
        # 填充数据
        for idx, invoice in enumerate(sorted_data, 1):
            row = [
                idx,
                invoice.get('buyer_name', ''),
                invoice.get('invoice_date', ''),
                invoice.get('invoice_type', ''),
                invoice.get('invoice_number', ''),
                invoice.get('amount', 0),
                invoice.get('tax', 0),
                invoice.get('amount', 0) + invoice.get('tax', 0)
            ]
            ws.append(row)
    
    def _create_summary_sheet(self, data: List[Dict]):
        '''创建汇总工作表'''
        ws = self.workbook.create_sheet('月度汇总')
        
        # 按月份分组统计
        monthly_stats = {}
        for invoice in data:
            date_str = invoice.get('invoice_date', '')
            if date_str:
                month = date_str[:7]  # 提取YYYY-MM
                if month not in monthly_stats:
                    monthly_stats[month] = {'count': 0, 'amount': 0, 'tax': 0}
                monthly_stats[month]['count'] += 1
                monthly_stats[month]['amount'] += invoice.get('amount', 0)
                monthly_stats[month]['tax'] += invoice.get('tax', 0)
        
        # 设置表头
        headers = ['月份', '发票数量', '总金额', '总税额', '价税合计']
        ws.append(headers)
        
        # 填充汇总数据
        for month in sorted(monthly_stats.keys()):
            stats = monthly_stats[month]
            row = [
                month,
                stats['count'],
                stats['amount'],
                stats['tax'],
                stats['amount'] + stats['tax']
            ]
            ws.append(row)